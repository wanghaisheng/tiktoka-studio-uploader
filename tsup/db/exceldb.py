import openpyxl


class ExcelDB:
    def __init__(self, file_path):
        self.file_path = file_path

    def find(self, sheet_name, limit=0, to_json=False, conver_col=True):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        data = list(rows)
        wb.close()

        if limit > 0:
            data = data[:limit]

        if to_json:
            columns = data[0]
            data = [dict(zip(columns, row)) for row in data[1:]]

        return data

    def add(self, sheet_name, data):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb[sheet_name]
        sheet.append(data)
        wb.save(self.file_path)
        wb.close()

    def update(self, sheet_name, data, condition):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        columns = next(rows)
        condition_column = condition.split("=")[0].strip()
        condition_value = condition.split("=")[1].strip()

        for row in rows:
            row_dict = dict(zip(columns, row))
            if row_dict.get(condition_column) == condition_value:
                for key, value in data.items():
                    column_index = columns.index(key)
                    sheet.cell(row=row[0], column=column_index + 1, value=value)
                break

        wb.save(self.file_path)
        wb.close()

    def delete(self, sheet_name, condition):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        columns = next(rows)
        condition_column = condition.split("=")[0].strip()
        condition_value = condition.split("=")[1].strip()

        for row in rows:
            row_dict = dict(zip(columns, row))
            if row_dict.get(condition_column) == condition_value:
                sheet.delete_rows(row[0])
                break

        wb.save(self.file_path)
        wb.close()


# excel_db = ExcelDB("path/to/excel/file.xlsx")
# data = excel_db.find("Sheet1")
# print(data)

# new_data = {'id': 2, 'name': 'Jane Doe', 'age': 25}
# excel_db.add("Sheet1", new_data)

# update_data = {'age': 26}
# excel_db.update("Sheet1", update_data, "id=1")

# excel_db.delete("Sheet1", "id=2")
